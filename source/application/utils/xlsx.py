import asyncio
from io import BytesIO
from typing import Tuple

from marshmallow import ValidationError
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.styles.numbers import BUILTIN_FORMATS, FORMAT_DATE_YYYYMMDD2
from openpyxl.utils import get_column_letter

from source.application.utils.hash import hash_payload
from source.constants import BULK_INSERT_LIMIT, LeadSourceType
from source.logger import get_logger
from source.models.lead import Lead
from source.schemas.lead import LeadXLSXSchema

FIRST_SHEET_INDEX = 0
ZERO_CELL = 0

HEADER_FIELDS = {
    '№ Аукциона': BUILTIN_FORMATS[49],
    'Дата': FORMAT_DATE_YYYYMMDD2,
    'Сумма БГ': BUILTIN_FORMATS[4],
    'Название': BUILTIN_FORMATS[49],
    'ИНН': BUILTIN_FORMATS[49],
    'E-mail': BUILTIN_FORMATS[49],
    'Конт. Лицо': BUILTIN_FORMATS[49],
    'Описание': BUILTIN_FORMATS[49],
    'Телефон': BUILTIN_FORMATS[49],
    'ФЗ': BUILTIN_FORMATS[49],
    'Регион': BUILTIN_FORMATS[49],
    'Наименование базы': BUILTIN_FORMATS[49],
    }

HEADER_FONT_SIZE = 12
HEADER_FONT = Font(name='Arial', size=HEADER_FONT_SIZE, bold=True)
FIRST_COLUMN = 1
FIRST_ROW = 1
ADJUST_CONST = 2  # experimental(could be changed) additional val to see the whole text in the cell
ADJUST_COEFFICIENT = HEADER_FONT_SIZE / 10

HEADER_BORDER = Border(left=Side(border_style=BORDER_THIN,
                                 color='FF000000'),
                       right=Side(border_style=BORDER_THIN,
                                  color='FF000000'),
                       top=Side(border_style=BORDER_THIN,
                                color='FF000000'),
                       bottom=Side(border_style=BORDER_THIN,
                                   color='FF000000'), )

XLSX_UPLOAD_ERROR_LIMIT = 100

log = get_logger('xlsx')


def load_row(row: tuple) -> dict:
    schema = LeadXLSXSchema()
    fields = schema.declared_fields.copy()

    for i, item in enumerate(fields):
        fields[item] = str(row[i].value) if row[i].value else None
    data = schema.load(fields)

    return data


def charge_error(error: dict, row_number: int = None):
    """Append additional info to error like ru_name of field and row number"""

    # new_error = error.copy()
    new_error = {}
    if isinstance(error, dict):
        for item, val in error.items():
            new_error[item] = {
                'ru_name': LeadXLSXSchema().declared_fields[item].metadata.get('description'),
                'row_number': row_number,
                'errors': val
                }
        return new_error


async def parse_xlsx(file: BytesIO, skip_header=True) -> Tuple[int, int, list]:
    """

    :param file: BytesIO: xlsx file
    :param skip_header: bool: do we need skip header
    :return: list of errors
    """
    errors = []
    uploaded_leads = 0
    to_insert_list = []
    wb = load_workbook(filename=file, data_only=True)
    ws = wb.worksheets[FIRST_SHEET_INDEX]
    for row in ws.iter_rows(min_row=FIRST_ROW + 1 if skip_header else FIRST_ROW,
                            min_col=FIRST_COLUMN,
                            max_col=len(HEADER_FIELDS)
                            ):
        try:
            data = load_row(row)
        except ValidationError as e:
            # todo: add rownumber to error
            log.info(f'Error occurred during xlsx row validation {e}')
            # to know which row it doesn't metter which Cell to take => take zero one
            err = charge_error(e.messages, row[ZERO_CELL].row)
            errors.append(err)
            if len(errors) >= XLSX_UPLOAD_ERROR_LIMIT:
                log.error(f'Too many erros {len(errors)} while uploading xlsx to db, break')
                break
            continue

        data['source_id'] = LeadSourceType.XLSX.value
        lead_hash = hash_payload(data)
        data['lead_hash'] = lead_hash

        to_insert_list.append(data)
        if len(to_insert_list) >= BULK_INSERT_LIMIT:
            await Lead.insert().gino.all(to_insert_list)
            uploaded_leads += len(to_insert_list)
            to_insert_list = []

    # if less then BULK_INSERT_LIMIT, and for last chunk
    await Lead.insert().gino.all(to_insert_list)
    uploaded_leads += len(to_insert_list)

    return uploaded_leads, len(errors), errors


def generate_xlsx(name: str = 'example.xlsx') -> BytesIO:
    wb = Workbook()
    ws = wb.active
    for i, (item, val) in enumerate(HEADER_FIELDS.items()):
        i += FIRST_COLUMN
        col_letter = get_column_letter(i)
        cell_name = col_letter + str(FIRST_COLUMN)

        ws[cell_name].font = HEADER_FONT  # can use ws(row=1, column=i)
        ws[cell_name].border = HEADER_BORDER
        col_dim = ws.column_dimensions[col_letter]
        col_dim.width = (len(item) + ADJUST_CONST) * ADJUST_COEFFICIENT
        col_dim.number_format = val
        ws[cell_name] = item

    output = BytesIO()
    wb.save(output)
    output.name = name
    return output


# inline tests
async def amain():
    from fwork.common.db.postgres.conn_async import db
    from fwork.common.db.postgres.settings import DSN

    await db.set_bind(DSN)
    name = 'test.xlsx'
    # generate_xlsx(name)
    name_to_load = '/Users/dmitriy/projects/f_work/lead_service/source/application/utils/example.xlsx'

    with open(name_to_load, 'rb') as file:
        leads, errors_count, errors = await parse_xlsx(file=BytesIO(file.read()))
    print(leads)
    print(errors_count)
    print(errors)


if __name__ == '__main__':
    asyncio.run(amain())
